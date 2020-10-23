# -*- coding: utf-8 -*-
"""
Created on Tue Jan 23 16:34:17 2018

@author: lab
"""

import os
import time
import openpyxl
import sys 
import copy
import json

import networkx as nx
import matplotlib.pyplot as plt

#import sys
#reload(sys)
#sys.setdefaultencoding('utf8')

RESULT_EXCEL = "result.xlsx"

def e2e_write_to_excel(index, data):
    row_0 = [u'端->端',u'路由',u'发包',u'收包',u'发包序号',u'收包序号',u'净收包数',u'交付率',u'净交付率',u'吞吐量',u'端延时']  
    exist_file = openpyxl.load_workbook(RESULT_EXCEL)
   
    table = exist_file.create_sheet()
    table.title = u"路由层"+str(index)
    
    for i in range(len(row_0)):
        table.cell(row=1, column=i+1).value = row_0[i]
    
    for row ,item in enumerate(data):
    #	print "row: ", row, " item: ", item
        for i in range(len(row_0)):
            table.cell(row=row+2, column=i+1).value = item[i]
    exist_file.save(RESULT_EXCEL)
    

def p2p_write_to_excel(index, data):
    row_0 = [u'点->点',u'跳延时',u'吞吐量', u'总发包',u'总收包',u'毛交付率',\
    u'发D包序号（可重复）', u'收D包序号（可重复）', u'总发D包数', u'总收D包数', u'重传D包数', u'净发D包数', u'净收D包数', u'重传率', u'D包毛交付率', u'D包净交付率',\
    u'ACK发包序号（可重复）',u'ACK收包序号（可重复）', u'总发ACK数', u'总收ACK数',u'ACK交付率', \
    u'B发包数', u'B收包数',u'B包交付率']  
    exist_file = openpyxl.load_workbook(RESULT_EXCEL)
   
    table = exist_file.create_sheet()
    table.title = u"MAC层"+str(index)
    for i in range(len(row_0)):
        table.cell(row=1, column=i+1).value = row_0[i]
    
    for row ,item in enumerate(data):
        for i in range(len(row_0)):
            table.cell(row=row+2, column=i+1).value = item[i]
    exist_file.save(RESULT_EXCEL)

"""
将总体网络性能信息写入Excel文件
"""
def performance_write_to_excel(data, start_pos, createflag=True):
    """
        performance_data = {"throughPut":"%.2f" % e2e_throughPut, 
                        "e2eSendNum": delivery_info['route_send_data_num'],
                        "e2eRecvNum": delivery_info['route_recv_data_num'],
                        "e2eDeliRate": delivery_info['route_deli_rate'], 
                        "e2eDropRate": delivery_info['route_drop_rate'],  
                        "aveDelay": "%.2f" % aveDelay + "s", 
                        "starttime": starttime, "endtime": endtime, "runtime": runtime}
    """
   
    first_row_0 = [u'吞吐量:', u'总发包数:', u'总收包数:', u'传送率:', u'丢包率:', u'平均传输时延:', u'节点开始发送数据时间:', u'所有节点停止发送接收数据时间：', u'网络工作时间：']
    dict_index = ["throughPut", "e2eSendNum", "e2eRecvNum", "e2eDeliRate", "e2eDropRate",\
                        "aveDelay", "starttime", "endtime", "runtime"]
    if createflag:
        f = openpyxl.Workbook()
        first_sheet = f.active
        for i in range(len(first_row_0)):
            first_sheet.cell(row=1, column=i+1).value = first_row_0[i]

        f.save(RESULT_EXCEL)

    
    exist_file = openpyxl.load_workbook(RESULT_EXCEL)
    first_sheet =exist_file.worksheets[0]
    for i in range(len(data)):
        first_sheet.cell(row=start_pos[0], column=i+start_pos[1]).value = data[dict_index[i]]
    exist_file.save(RESULT_EXCEL)

    
if __name__ == '__main__':
    UWAN = UWANPerformance()
    UWAN.read_log_file()
    
    '''
    整个网络运行时间总体性能，写入excel表第一页
    '''
    UWAN.performance(UWAN.loglist)
    with open("result.json", "w") as f:
    	json.dump(UWAN.json_data, f)
	print "json dump finish..."
    performance_write_to_excel(UWAN.performance_data, [2,1], createflag=False)

#    '''
#    路由有间隔更新的情况，若没有可将下面一段注释掉
#    1.将某时间段的整体性能写入excel第一页
#    2.将某时间段端到端的情况统计写入excel的sheet
#    '''

    """
    split_loglist = UWAN.split_by_route()
    UWAN.startTime =0
    UWAN.endTime = 0
    UWAN.row_count = 3
    UWAN.sheet_num = 1
    """

"""
    for log in split_loglist:  
        routelist = UWAN.find_route_table(log)
#        print "routelist2", routelist
        if len(routelist) > 0:
            print "====================="
            UWAN.performance(log)
	

            performance_write_to_excel(UWAN.performance_data, [UWAN.row_count,1], createflag=True)
            UWAN.row_count += 1
            UWAN.sheet_num += 1
            path_list = UWAN.draw_topo(routelist, draw=False)
            sheet_route_data = UWAN.end2end(log, path_list)
            sheet_mac_data = UWAN.point2point(log)
            
            UWAN.route_data.append(sheet_route_data)
            UWAN.mac_data.append(sheet_mac_data)
           
    for index, each_route_data in enumerate(UWAN.route_data):
        e2e_write_to_excel(index+1, each_route_data)
    for index, each_mac_data in enumerate(UWAN.mac_data):
        p2p_write_to_excel(index+1, each_mac_data)

"""
